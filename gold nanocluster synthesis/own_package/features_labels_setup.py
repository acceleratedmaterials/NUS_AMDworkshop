import numpy as np
import numpy.random as rng
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import StratifiedKFold, KFold, LeaveOneOut
from keras.utils import to_categorical
import pickle
import os
import pathlib
import warnings
import random
# Own script
from .others import print_array_to_excel
from .NGF.preprocessing import tensorise_smiles

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.expand_frame_repr', False)


def read_reaction_data(loader_excel_file='./excel/data_loader.xlsx', mode='c', save_mode=False, print_mode=False):
    '''
    Used to read reaction data from excel file containing reaction data. Must split the workbook into 4 sheets
    df : sheet containing the headers and values for all reaction data
    features_c: sheet containing continuous features values. Each row is one example with each col being diff. features
    features_d: sheet containing categorical features. As of now, the code can only handle 1 categorical feature. :(
    :param loader_excel_file: excel file name
    :param print_mode: Print results
    :return: fl class object
    '''
    try:
        df = pd.read_excel(loader_excel_file, sheet_name='df')
    except FileNotFoundError:
        try:
            # Check for xlsx file with the same name but with file extension missing
            if loader_excel_file[-5:] != '.xlsx':
                # Must create a temp name because there is another check after this
                loader_excel_file1 = loader_excel_file + '.xlsx'
                df = pd.read_excel(loader_excel_file1, sheet_name='df')
            else:
                # Means that even with .xlsx at the back, file is not found, so raise an error.
                raise FileNotFoundError
        except FileNotFoundError:
            # Check for xlmx file instead
            if loader_excel_file[-5:] != '.xlsm':
                loader_excel_file = loader_excel_file + '.xlsm'
            df = pd.read_excel(loader_excel_file, sheet_name='df')
        else:
            # If first check succeeds, rename excel file name with the temp file name
            loader_excel_file = loader_excel_file1
    features_c = pd.read_excel(loader_excel_file, sheet_name='features_c')
    features_c_values = features_c.values  # ndarray
    features_c_names = features_c.columns.values  # ndarray
    # There should be one column in the excel sheet for labels only!
    if mode == 'c':
        labels = pd.read_excel(loader_excel_file, sheet_name='labels').values.flatten()
    elif mode == 'r':
        labels = pd.read_excel(loader_excel_file, sheet_name='labels').values
    else:
        raise TypeError('mode input should be either c or r')
    fl = Features_labels(features_c_values, labels, features_c_names=features_c_names, mode=mode, save_mode=save_mode)
    if print_mode:
        print(df)
    return fl


def read_reaction_data_smiles(loader_excel_file='./excel/data_loader.xlsx', mode='c', save_mode=False,
                              print_mode=False):
    '''
    Used to read reaction data from excel file containing reaction data. Must split the workbook into 3 sheets
    df : sheet containing the headers and values for all reaction data
    features_c: sheet containing continuous features values. Each row is one example with each col being diff. features
    features_d: sheet containing smiles information.
    :param loader_excel_file: excel file name
    :param mode: To set as classification or regression mode with the string 'c' or 'r' respectively
    :param save_mode: Whether to save the fl object
    :param print_mode: Print results
    :return: fl class object
    '''
    try:
        df = pd.read_excel(loader_excel_file, sheet_name='df')
    except FileNotFoundError:
        try:
            # Check for xlsx file with the same name but with file extension missing
            if loader_excel_file[-5:] != '.xlsx':
                # Must create a temp name because there is another check after this
                loader_excel_file1 = loader_excel_file + '.xlsx'
                df = pd.read_excel(loader_excel_file1, sheet_name='df')
            else:
                # Means that even with .xlsx at the back, file is not found, so raise an error.
                raise FileNotFoundError
        except FileNotFoundError:
            # Check for xlmx file instead
            if loader_excel_file[-5:] != '.xlsm':
                loader_excel_file = loader_excel_file + '.xlsm'
            df = pd.read_excel(loader_excel_file, sheet_name='df')
        else:
            # If first check succeeds, rename excel file name with the temp file name
            loader_excel_file = loader_excel_file1
    # There should be one column in the excel sheet for labels only!
    if mode == 'c':
        labels = pd.read_excel(loader_excel_file, sheet_name='labels').values.flatten()
    elif mode == 'r':
        labels = pd.read_excel(loader_excel_file, sheet_name='labels').values
    else:
        raise TypeError('mode input should be either c or r')
    # Contains continous features. 1st dim is example, 2nd dim is the features
    features_c = pd.read_excel(loader_excel_file, sheet_name='features_c')
    features_c_values = features_c.values  # ndarray
    features_c_names = features_c.columns.values  # ndarray
    # Contains smiles. (example, no. of molecule inputs) for example, (10 examples, 3 molecules)
    features_d = pd.read_excel(loader_excel_file, sheet_name='features_d')
    features_d_values = features_d.values
    features_d_names = features_d.columns.values
    # Contains atom, bond, edge tensor for each molecule. (no. molecules, 3, ndarray).
    # ndarray is (examples, ...)
    features_d_store = []
    for single_fd in features_d_values.T:
        features_d_store.append(tensorise_smiles(single_fd))

    fl = Features_labels_smiles(features_c_values, features_d_store,
                                labels, smiles=features_d.values, features_c_names=features_c_names,
                                features_d_names=features_d_names, mode=mode, scaler=None, num_classes=None,
                                save_mode=save_mode)
    if print_mode:
        print(df)
    return fl


class Features_labels:
    def __init__(self, features_c, labels, mode='c', features_c_names=None, scaler=None, num_classes=None,
                 idx=None,
                 save_name='fl', save_mode=False):
        """
        Creates fl class with a lot useful attributes
        :param features_c: Continuous features. Np array, no. of examples x continous features
        :param labels: Labels as np array, no. of examples x 1
        :param mode: c for classification mode, r for regression mode
        :param scaler: Scaler to transform features c. If given, use given MinMax scaler from sklearn,
        else create scaler based on given features c.
        :param save_name: If save mode is on, will be the name of the saved fl class obj
        :param save_mode: To save or not.
        """
        ###################
        def features_to_listedtuple(features, targets):
            '''
            Converts features from np array to list of sorted tuples based on labels
            :param features: features to be sorted
            :param targets: labels
            :return: list of sorted tuples. eg: [(0, features nxm), (1, features nxm)]
            '''
            dic = {}
            for feature, target in zip(np.ndarray.tolist(features), targets):
                if target in dic:  # Create new class tuple in the dic
                    dic[target].append(feature)
                else:  # If class already exists, append new features into that class
                    dic[target] = [feature]
            for target in dic:  # Convert list back to ndarray
                dic[target] = np.array(dic[target])
            # Convert from dictionary to list of tuple
            return sorted(dic.items())
        
        ####################
        
        self.mode = mode
        self.features_c_names = features_c_names

        if isinstance(idx, np.ndarray):
            self.idx = idx
        else:
            self.idx = np.arange(features_c.shape[0])

        if mode == 'c':  # Means classification mode
            # Setting up features
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Without _a at the back means it is the listed tuple data type.
            self.features_c = features_to_listedtuple(features_c, labels)
            # Normalizing continuous features
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
            self.features_c_norm = features_to_listedtuple(self.features_c_norm_a, labels)
            # Setting up labels
            self.labels = labels
            if num_classes is None:
                _, count = np.unique(labels, return_counts=True)
                self.n_classes = len(count)
            else:
                self.n_classes = num_classes
            self.labels_hot = to_categorical(labels, num_classes=self.n_classes)
            # List containing number of examples per class
            self.count_per_class = [category[1].shape[0] for category in self.features_c]
            # Storing dimensions
            self.features_c_dim = features_c.shape[1]
        elif mode == 'r':  # Means regression mode
            # Setting up features
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Normalizing continuous features
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
            # Setting up labels
            self.labels = labels
            # Storing dimensions
            self.features_c_dim = features_c.shape[1]
            if len(labels.shape) == 2:
                self.labels_dim = labels.shape[1]
            else:
                self.labels_dim = 1

        # Saving
        if save_mode:
            file_path = open('./save/features_labels/' + save_name + '.obj', 'wb')
            pickle.dump(self, file_path)

    def generate_random_examples(self, numel):
        gen_features_c_norm_a = rng.random_sample((numel, self.features_c_dim))
        gen_features_c_a = self.scaler.inverse_transform(gen_features_c_norm_a)

        # Creating dic for SNN prediction
        gen_dic = {}
        gen_dic = dict(
            zip(('gen_features_c_a', 'gen_features_c_norm_a'), (gen_features_c_a, gen_features_c_norm_a)))
        return gen_dic

    def generate_random_subset(self, subset_split, save_fl=False, save_to_excel=False,
                               loader_excel_file='./excel/data_loader.xlsx'):
        '''
        Split main data set of examples into subsets containing the desired number of classes. For example,
        subset_split = [3,3,3] means make a subset with 3 class 0s examples, 3 class 1s, 3 class 2s
        :param subset_split: List with number of elements equal to total number of class
        :param loader_excel_file: Name of data loader excel file to be open and to save stuff to
        :return:
        '''
        # Setting up subset examples
        for category in range(self.n_classes):
            try:
                # Subset
                n_examples = self.features_c_norm[category][1].shape[0]  # Get no. of examples for current class
                idx = rng.choice(n_examples, size=(subset_split[category],), replace=False)
                features_c_a = self.features_c[category][1][idx, :]  # Choose number of examples for current class
                labels = np.repeat(category, subset_split[category])  # Set the labels for the selected examples
                # Inverse subset. Deleting the examples chosen for the subset
                i_features_c_a = np.delete(self.features_c[category][1], idx, axis=0)
                i_labels = np.repeat(category, n_examples - subset_split[category])
                if category == 0:  # Means first loop and the array is not formed yet
                    features_c_a_store = features_c_a
                    labels_store = labels
                    i_features_c_a_store = i_features_c_a
                    i_labels_store = i_labels
                else:
                    features_c_a_store = np.concatenate((features_c_a_store, features_c_a), axis=0)
                    labels_store = np.concatenate((labels_store, labels))
                    i_features_c_a_store = np.concatenate((i_features_c_a_store, i_features_c_a), axis=0)
                    i_labels_store = np.concatenate((i_labels_store, i_labels))
            except:
                if len(subset_split) != self.n_classes:
                    raise Exception('Subset split does not have same number of elements as the total number of class!'
                                    'Make sure that they are the same.')
                continue
        ss_fl = Features_labels(features_c_a_store, labels_store, save_name='ss_fl', save_mode=save_fl)
        # Setting up inverse subset examples. Note: Must use ss_scaler to ensure that when doing the prediction, the
        # inverse subset is scaled correctly and the same as when the model is trained with the subset examples
        ss_scaler = ss_fl.scaler
        i_ss_fl = Features_labels(i_features_c_a_store, i_labels_store, scaler=ss_scaler,
                                  save_name='i_ss_fl', save_mode=save_fl)
        # Excel writing part
        if save_to_excel:
            wb = load_workbook(loader_excel_file)
            # Setting up subset features and labels sheet
            sheet_name_store = ['ss_features_c', 'ss_labels', 'i_ss_features_c', 'i_ss_labels']
            ss_store = [features_c_a_store, labels_store, i_features_c_a_store, i_labels_store]
            axis_store = [2, 0, 2, 0]  # Because feature_c is 2D while labels is col vector, so order of axis is 2,0,2,0
            for cnt, sheet_name in enumerate(sheet_name_store):
                if sheet_name in wb.sheetnames:
                    # If temp sheet exists, remove it to create a new one. Else skip this.
                    idx = wb.sheetnames.index(sheet_name)  # index of temp sheet
                    wb.remove(wb.worksheets[idx])  # remove temp
                    wb.create_sheet(sheet_name, idx)  # create an empty sheet using old index
                else:
                    wb.create_sheet(sheet_name)  # Create the new sheet
                # Print array to the correct worksheet
                print_array_to_excel(ss_store[cnt], (2, 1), wb[sheet_name_store[cnt]], axis=axis_store[cnt])
            wb.save(loader_excel_file)
            wb.close()
        return ss_fl, i_ss_fl

    def create_skf(self, k_folds):
        '''
        Create list of tuples containing (fl_train,fl_val) fl objects for stratified k fold cross validation
        :param k_folds: Number of folds
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        skf = StratifiedKFold(n_splits=k_folds, shuffle=True)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels(xtrain, ytrain, scaler=self.scaler, num_classes=self.n_classes, save_mode=False),
                 Features_labels(xval, yval, scaler=self.scaler, num_classes=self.n_classes, save_mode=False))
            )
        return fl_store

    def create_kf(self, k_folds, shuffle=True):
        '''
        Almost the same as skf except can work for regression labels and folds are not stratified.
        Create list of tuples containing (fl_train,fl_val) fl objects for k fold cross validation
        :param k_folds: Number of folds
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=shuffle)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xval_idx = self.idx[val_indices]
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels(xtrain, ytrain, scaler=self.scaler, num_classes=self.n_classes, save_mode=False),
                 Features_labels(xval, yval, idx=xval_idx, scaler=self.scaler, num_classes=self.n_classes,
                                 save_mode=False))
            )
        return fl_store

    def create_loocv(self):
        '''
        Create list of tuples containing (fl_train,fl_val) fl objects for leave one out cross validation
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        loocv = LeaveOneOut()
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(loocv.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels(xtrain, ytrain, scaler=self.scaler, num_classes=self.n_classes, save_mode=False),
                 Features_labels(xval, yval, scaler=self.scaler, num_classes=self.n_classes, save_mode=False))
            )
        return fl_store

    def write_data_to_excel(self, loader_excel_file='./excel/data_loader.xlsx'):
        # Excel writing part
        wb = load_workbook(loader_excel_file)
        # Setting up subset features and labels sheet
        sheet_name_store = ['temp_features_c', 'temp_labels']
        ss_store = [self.features_c_a, self.labels]
        axis_store = [2, 0]  # Because feature_c is 2D while labels is col vector, so order of axis is 2,0
        for cnt, sheet_name in enumerate(sheet_name_store):
            if sheet_name in wb.sheetnames:
                # If temp sheet exists, remove it to create a new one. Else skip this.
                idx = wb.sheetnames.index(sheet_name)  # index of temp sheet
                wb.remove(wb.worksheets[idx])  # remove temp
                wb.create_sheet(sheet_name, idx)  # create an empty sheet using old index
            else:
                wb.create_sheet(sheet_name)  # Create the new sheet
            # Print array to the correct worksheet
            print_array_to_excel(ss_store[cnt], (2, 1), wb[sheet_name_store[cnt]], axis=axis_store[cnt])
        wb.save(loader_excel_file)
        wb.close()

    def create_subsetsplit(self, subsetsplit):
        '''
        Split the main fl into multiple different fl depending on the given subsetsplit.
        Example:
        [5] ==> split into [0:5] and [5:]
        [3,5] ==> split into [0:3], [3:5], and [5:]
        :param subsetsplit: list of numbers or a single int value.
        :return: fl_store, a list containing the fl splits.
        '''

        if not isinstance(subsetsplit, list):
            assert isinstance(subsetsplit, int), 'subsetsplit must be either a list of len>0 or single int value'
            subsetsplit = list(subsetsplit)

        if subsetsplit[-1] < self.count:
            # Add the last index to slice to the end of subsetsplit
            subsetsplit.append(self.count)
        elif subsetsplit[-1] == self.count:
            warnings.warn('subsetsplit last indice should not be the last index of the total fl. '
                          'Treating the last indice as the last index of the subset.')
        else:
            raise TypeError('subsetsplit last element indice is greater than original fl dimension. '
                            'Choose last indice to be < original f1 dimension')
        fl_store = []
        if subsetsplit[0] == 0:
            # First element should not be 0. But if it is, just remove it and continue.
            subsetsplit.pop(0)
        left_idx = 0
        total_indices = np.arange(self.count)
        for idx in subsetsplit[:]:
            right_idx = subsetsplit.pop(0)
            indices = total_indices[left_idx:right_idx]
            features_c = self.features_c_a[indices]
            labels = self.labels[indices]

            fl_store.append(
                Features_labels(features_c, labels, features_c_names=self.features_c_names,
                                scaler=self.scaler, num_classes=self.n_classes, save_mode=False)
            )

            left_idx = right_idx

        return fl_store
    
    
    def random_split(self, num):
        if num >= self.count:
            return self, None
        
        
        indices = random.sample(list(np.arange(self.count)), num) 
        for i in range(len(indices)):
            self.features_c_a[i] , self.features_c_a[indices[i]] =self.features_c_a[indices[i]] ,self.features_c_a[i]
            self.labels[i] , self.labels[indices[i]] =self.labels[indices[i]] ,self.labels[i]
        return self.create_subsetsplit([0,num])
            
        

    def print_all(self):
        print('features_c : \n{} \n'
              'features_c_norm : \n{} \n'
              'features_c_a : \n{}\n'
              'labels : \n{} \n'
              'dim of c = {} '.format(self.features_c,
                                      self.features_c_norm,
                                      self.features_c_a,
                                      self.labels,
                                      self.features_c_dim, ))


class Features_labels_smiles:
    def __init__(self, features_c, features_d, labels, smiles=None, features_c_names=None, features_d_names=None,
                 idx=None, mode='c',
                 scaler=None,
                 num_classes=None, save_name='fl_smiles', save_mode=False):
        """
        Creates fl class with a lot useful attributes
        :param features_c: Continuous features. Np array, no. of examples x continous features
        :param features_d: SMILES data for discrete molecules in the order of solvent, ligand, RA
        :param labels: Labels as np array, no. of examples x 1
        :param mode: c for classification mode, r for regression mode, p for prediction
        :param scaler: Scaler to transform features c. If given, use given MinMax scaler from sklearn,
        else create scaler based on given features c.
        :param save_name: If save mode is on, will be the name of the saved fl class obj
        :param save_mode: To save or not.
        """

        def features_to_listedtuple(features, targets):
            '''
            Converts features from np array to list of sorted tuples based on labels
            :param features: features to be sorted
            :param targets: labels
            :return: list of sorted tuples. eg: [(0, features nxm), (1, features nxm)]
            '''
            dic = {}
            try:
                for feature, target in zip(np.ndarray.tolist(features), targets):
                    if target in dic:  # Create new class tuple in the dic
                        dic[target].append(feature)
                    else:  # If class already exists, append new features into that class
                        dic[target] = [feature]
            except TypeError:
                for feature, target in zip(features, targets):
                    if target in dic:  # Create new class tuple in the dic
                        dic[target].append(feature)
                    else:  # If class already exists, append new features into that class
                        dic[target] = [feature]
            for target in dic:  # Convert list back to ndarray
                dic[target] = np.array(dic[target])
            # Convert from dictionary to list of tuple
            return sorted(dic.items())

        # Checking if all inputs are ndarray
        if not isinstance(features_c, np.ndarray):
            features_c = np.array(features_c)
        if not isinstance(labels, np.ndarray):
            labels = np.array(labels)

        self.mode = mode
        self.features_c_names = features_c_names
        self.features_d_names = features_d_names

        if isinstance(smiles, np.ndarray):
            self.smiles = smiles
        else:
            self.smiles = None
        if isinstance(idx, np.ndarray):
            self.idx = idx
        else:
            self.idx = np.arange(features_c.shape[0])

        if mode == 'c':  # Means classification mode
            # Setting up features_c
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Without _a at the back means it is the listed tuple data type.
            self.features_c = features_to_listedtuple(features_c, labels)
            # Normalizing continuous features, self.labels
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
            self.features_c_norm = features_to_listedtuple(self.features_c_norm_a, labels)
            # Storing dimensions
            self.features_c_dim = features_c.shape[1]

            # Setting up features_d. Features containing SMILES for solvent, ligand, RA
            # Shape (no. of molecules, 3 tensors for a,b,e)
            self.features_d_a = features_d
            self.features_d_count = len(features_d)
            self.features_d = []
            # features_d shape: [no. molecules, tuple of 3 for a_b_e tensors, listed_tuples of (no. class, tensor))
            for i in range(self.features_d_count):
                single_fd = list(map(lambda x: features_to_listedtuple(x, labels), self.features_d_a[i]))
                self.features_d.append(single_fd)

            # Setting up labels
            self.labels = labels
            if num_classes is None:
                _, count = np.unique(labels, return_counts=True)
                self.n_classes = len(count)
            else:
                self.n_classes = num_classes
            self.labels_hot = to_categorical(labels, num_classes=self.n_classes)
            # List containing number of examples per class
            self.count_per_class = [category[1].shape[0] for category in self.features_c]

        elif mode == 'r':  # Means regression mode
            # Setting up features
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Normalizing continuous features
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
            # Storing dimensions
            self.features_c_dim = features_c.shape[1]
            # Setting up features_d. Features containing SMILES for solvent, ligand, RA
            # Contains atom, bond, edge tensor for each molecule. (no. molecules, 3, ndarray).
            # ndarray is (examples, ...)
            self.features_d_a = features_d
            self.features_d_count = len(features_d)
            self.features_d = []
            for i in range(self.features_d_count):
                single_fd = list(map(lambda x: features_to_listedtuple(x, labels), self.features_d_a[i]))
                self.features_d.append(single_fd)

            # Setting up labels
            self.labels = labels
            if len(labels.shape) == 2:
                self.labels_dim = labels.shape[1]
            else:
                self.labels_dim = 1

        elif mode == 'p':  # Prediction mode. labels==None
            # Setting up features
            self.count = features_c.shape[0]
            self.features_c_count = features_c.shape[1]
            # _a at the back means it is a ndarray type
            self.features_c_a = features_c
            # Normalizing continuous features
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.scaler.fit(features_c)  # Setting up scaler
            self.features_c_norm_a = self.scaler.transform(features_c)  # Normalizing features_c
            # Storing dimensions
            self.features_c_dim = features_c.shape[1]

            # Setting up features_d. Features containing SMILES for solvent, ligand, RA
            # 1,2,3 represents the molecule for the solvent, ligand, RA. x,y,z for atom, bond, edge tensor
            # Setting up features_d. Features containing SMILES for solvent, ligand, RA
            self.features_d_a = features_d
            self.features_d_count = len(features_d)

        # Saving
        if save_mode:
            pathlib.Path('./save/features_labels/').mkdir(parents=True, exist_ok=True)
            file_path = open('./save/features_labels/' + save_name + '.obj', 'wb')
            pickle.dump(self, file_path)
            file_path.close()

    def create_loocv(self):
        '''
        Create list of tuples containing (fl_train,fl_val) fl objects for leave one out cross validation
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        loocv = LeaveOneOut()
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(loocv.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            xtrain_d = []
            xval_d = []
            xval_smiles = self.smiles[val_indices, :]
            xval_idx = self.idx[val_indices]

            for single_molecule in self.features_d_a:
                xtrain_d.append((single_molecule[0][train_indices, ...],
                                 single_molecule[1][train_indices, ...],
                                 single_molecule[2][train_indices, ...],))
                xval_d.append((single_molecule[0][val_indices, ...],
                               single_molecule[1][val_indices, ...],
                               single_molecule[2][val_indices, ...],))
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels_smiles(xtrain, xtrain_d, ytrain, smiles=None, mode=self.mode,
                                        features_c_names=self.features_c_names, features_d_names=self.features_d_names,
                                        scaler=self.scaler, num_classes=self.n_classes, save_mode=False),
                 Features_labels_smiles(xval, xval_d, yval, smiles=xval_smiles, idx=xval_idx, mode=self.mode,
                                        features_c_names=self.features_c_names, features_d_names=self.features_d_names,
                                        scaler=self.scaler, num_classes=self.n_classes, save_mode=False)
                 )
            )
        return fl_store

    def create_kf(self, k_folds=10, shuffle=True):
        '''
        Create list of tuples containing (fl_train,fl_val) fl objects for stratified k-fold cross validation
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=shuffle)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c_a, self.labels)):
            # Generate batches from indices
            xtrain, xval = self.features_c_a[train_indices], self.features_c_a[val_indices]
            xtrain_d = []
            xval_d = []
            xval_smiles = self.smiles[val_indices, :]
            xval_idx = self.idx[val_indices]

            for single_molecule in self.features_d_a:
                xtrain_d.append((single_molecule[0][train_indices, ...],
                                 single_molecule[1][train_indices, ...],
                                 single_molecule[2][train_indices, ...],))
                xval_d.append((single_molecule[0][val_indices, ...],
                               single_molecule[1][val_indices, ...],
                               single_molecule[2][val_indices, ...],))

            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels_smiles(xtrain, xtrain_d, ytrain, smiles=None, mode=self.mode,
                                        features_c_names=self.features_c_names, features_d_names=self.features_d_names,
                                        scaler=self.scaler, num_classes=self.n_classes, save_mode=False),
                 Features_labels_smiles(xval, xval_d, yval, smiles=xval_smiles, idx=xval_idx, mode=self.mode,
                                        features_c_names=self.features_c_names, features_d_names=self.features_d_names,
                                        scaler=self.scaler, num_classes=self.n_classes, save_mode=False)
                 )
            )
        return fl_store

    def create_subsetsplit(self, subsetsplit):
        '''
        Split the main fl into multiple different fl depending on the given subsetsplit.
        Example:
        [5] ==> split into [0:5] and [5:]
        [3,5] ==> split into [0:3], [3:5], and [5:]
        :param subsetsplit: list of numbers or a single int value.
        :return: fl_store, a list containing the fl splits.
        '''

        if not isinstance(subsetsplit, list):
            assert isinstance(subsetsplit, int), 'subsetsplit must be either a list of len>0 or single int value'
            subsetsplit = list(subsetsplit)

        if subsetsplit[-1] < self.count:
            # Add the last index to slice to the end of subsetsplit
            subsetsplit.append(self.count)
        elif subsetsplit[-1] == self.count:
            warnings.warn('subsetsplit last indice should not be the last index of the total fl. '
                          'Treating the last indice as the last index of the subset.')
        else:
            raise TypeError('subsetsplit last element indice is greater than original fl dimension. '
                            'Choose last indice to be < original f1 dimension')
        fl_store = []
        left_idx = 0
        total_indices = np.arange(self.count)
        for idx in subsetsplit[:]:
            right_idx = subsetsplit.pop(0)
            indices = total_indices[left_idx:right_idx]
            features_c = self.features_c_a[indices]
            features_d = []
            fl_smiles = self.smiles[indices]
            fl_idx = self.idx[indices]
            labels = self.labels[indices]

            for single_molecule in self.features_d_a:
                features_d.append((single_molecule[0][indices, ...],
                                   single_molecule[1][indices, ...],
                                   single_molecule[2][indices, ...],))

            fl_store.append(
                Features_labels_smiles(features_c, features_d, labels, smiles=fl_smiles, idx=fl_idx, mode=self.mode,
                                       features_c_names=self.features_c_names, features_d_names=self.features_d_names,
                                       scaler=self.scaler, num_classes=self.n_classes, save_mode=False)
            )

            left_idx = right_idx

        return fl_store
  
        
    
#    NEED SOME WORK HERE    
#    def random_split(self, num):
#        if num >= self.count:
#            return self, None
#        
#        
#        indices = random.sample(list(np.arange(self.count)), num) 
#        for i in range(len(indices)):
#            self.feature_c_a[i] , self.feature_c_a[indices[i]] =self.feature_c_a[indices[i]] ,self.feature_c_a[i]
#            self.smiles[i] , self.smiles[indices[i]] =self.smiles[indices[i]] ,self.smiles[i]
#            
#            
#        return self.create_subsplit([0,num])
    
    

    # Generate batches from smile indices. Is to generate random features_c for a given set of features_d
    def generate_examples(self, smiles_idx, numel=10, mode=0, **kwargs):
        '''
        2 types of method to generate examples to evaluate using a trained model.

        Mode 0.
        Choose a smiles_idx to indicate the type of solvent, ligand, RA molecule to use for all generated examples.
        Randomly generates features_c within min, max value of training data

        Mode 1.
        Choose a smiles_idx to indicate the type of solvent, ligand, RA molecule to use for all generated examples.
        Choose selected_c to choose the indices of the features_c columns to linspace.
        Choose selected_range to determine the min, max value for the cols selected in selected_c
        Choose numel, either int or list, to determine the number of permutation for each col.
        A meshgrid will be formed such that all possible combination of permutation of selected_c is given as a
        generated example.
        The un-selected features_c cols will be have the same value as the smiles_idx row example in the training
        data loader.

        :param smiles_idx: Int. Choose solvent, ligand, RA molecules to use for all examples for mode 0 or 1.
        :param numel: Int or list of len(selected_c). For mode 0, is to set total number of examples to generate.
        For mode 1, is to set number of examples for each linspace of a features_c col.
        Total_numel = product of all numel
        :param mode: Choose the mode type to use
        :param kwargs:
        For mode 1:
        **selected_c: (int or list) choose which cols of features_c to linspace.
        **selected_range: (list or nested list of shape len(selected_c) x 2) set min, max values for each col of
        generated features_c
        :return:
        '''

        assert self.smiles is not None, 'fl object does not contain smiles attribute. ' \
                                        'Ensure that smiles is given when initialising fl class'
        if mode == 0:
            # Select the molecules for solvent, ligand, RA by choosing smiles_idx
            # First slice for to choose atom, bond, or edge tensor. 2nd slice to choose the specific example data.
            # 3rd slice is to create new axis at 1st dim to repeat numel times to input together with gen_features_c
            features_d_a = []
            for single_molecule in self.features_d_a:
                features_d_a.append((np.repeat(single_molecule[0][smiles_idx, ...][None, ...], repeats=numel, axis=0),
                                     np.repeat(single_molecule[1][smiles_idx, ...][None, ...], repeats=numel, axis=0),
                                     np.repeat(single_molecule[2][smiles_idx, ...][None, ...], repeats=numel, axis=0),))

            # Creating SMILES array
            smiles = np.repeat(self.smiles[smiles_idx, :][None, :], repeats=numel, axis=0)

            # Randomly generate features_c
            gen_features_c_norm_a = rng.random_sample((numel, self.features_c_dim))
            gen_features_c_a = self.scaler.inverse_transform(gen_features_c_norm_a)

        elif mode == 1:
            try:
                selected_c = kwargs['selected_c']
                selected_range = kwargs['selected_range']
            except KeyError:
                raise KeyError('Mode 1 selected. selected_c and/ or selected_range is not given. '
                               'Give those two as a 1D and 2D nested list respectively')
            try:
                if isinstance(numel, int):
                    numel = [numel for _ in range(len(selected_c))]
                numel_final = np.prod(numel)
            except TypeError:
                if isinstance(selected_c, int):
                    if not any(isinstance(i, list) for i in selected_range):
                        selected_c = [selected_c]  # Change int to [X} list of 1 element only
                        selected_range = [selected_range]  # Change [X,Y] to [[X,Y]] nested list
                        numel_final = numel ** len(
                            selected_c)  # new numel which accounts for all permutation of selected c
                    else:
                        raise TypeError('Selected range has min, max value for more than 1 col of feature '
                                        'but only 1 col index of feature c is given in selected_c')
                else:
                    raise TypeError('Ensure both selected_c and selected_range are lists.')

            # Select the molecules for solvent, ligand, RA by choosing smiles_idx
            # First slice for to choose atom, bond, or edge tensor. 2nd slice to choose the specific example data.
            # 3rd slice is to create new axis at 1st dim to repeat numel times to input together with gen_features_c
            features_d_a = []
            for single_molecule in self.features_d_a:
                a= single_molecule[0][smiles_idx, ...][None, ...]
                features_d_a.append(
                    (np.repeat(a, repeats=numel_final, axis=0),
                     np.repeat(single_molecule[1][smiles_idx, ...][None, ...], repeats=numel_final, axis=0),
                     np.repeat(single_molecule[2][smiles_idx, ...][None, ...], repeats=numel_final, axis=0),))

            # Creating SMILES array
            smiles = np.repeat(self.smiles[smiles_idx, :][None, :], repeats=numel_final, axis=0)

            # Generating feature_c
            mask = np.ones(self.features_c_dim, dtype=np.bool)
            mask[selected_c] = 0
            gen_features_c_a = np.zeros((numel_final, self.features_c_dim))
            # to put into cols of inverse selection, repeat the features_c_a invese columns of selected smiles_idx row
            # for numel number of times to make shape (numel, mask)
            gen_features_c_a[:, mask] = np.repeat(self.features_c_a[smiles_idx, mask][None, :], repeats=numel_final,
                                                  axis=0)

            gen_store = []
            for single_numel, single_range, single_c in zip(numel, selected_range, selected_c):
                gen_single_features_c = np.linspace(start=single_range[0], stop=single_range[1],endpoint=True, num=single_numel)
                gen_store.append(gen_single_features_c)

            gen_selected_features_c = np.stack(np.meshgrid(*gen_store), axis=-1).reshape(-1, len(gen_store))
            gen_features_c_a[:, selected_c] = gen_selected_features_c
#        elif mode == 10:
            #rewriting the mode 1, for only selected_c =[**elements] and selected_range = [**[range]]
            #for memory saving purpose
            
            
            
        elif mode == 2:
            try:
                selected_c = kwargs['selected_c']
                selected_tolerance = kwargs['selected_tolerance']
            except KeyError:
                raise KeyError('Mode 2 selected. selected_c and/ or selected_tolerance is not given. '
                               'Give those two as a 1D and 2D nested list respectively')
            if isinstance(selected_c, int):
                # Check if selected_tolerance is int too.
                if isinstance(selected_tolerance, int):
                    # If so, convert both selected_c and selected_tolerance to list
                    selected_c = [selected_c]  # Change int to [X} list of 1 element only
                    selected_tolerance = [selected_tolerance]  # change int to [X] list of 1 element only
                else:
                    raise TypeError('selected_c is int but selected_tolerance is not int.'
                                    'Ensure both selected_c and selected_range are lists of same length,'
                                    ' or both int. ')

            sum_features_c_a = []
            sum_features_d_a = [[[] for _ in range(3)] for _ in range(self.features_d_count)]
            sum_smiles = []

            for idx in range(self.count):
                # Select the molecules for solvent, ligand, RA by choosing smiles_idx
                # First slice for to choose atom, bond, or edge tensor. 2nd slice to choose the specific example data.
                # 3rd slice is to create new axis at 1st dim to repeat numel times to input together with gen_features_c
                for molecular_idx, single_molecule in enumerate(self.features_d_a):
                    try:
                        sum_features_d_a[molecular_idx][0] = np.concatenate(
                            (sum_features_d_a[molecular_idx][0],
                             np.repeat(single_molecule[0][idx, ...][None, ...], repeats=numel, axis=0)),
                            axis=0)
                        sum_features_d_a[molecular_idx][1] = np.concatenate(
                            (sum_features_d_a[molecular_idx][1],
                             np.repeat(single_molecule[1][idx, ...][None, ...], repeats=numel, axis=0)),
                            axis=0)
                        sum_features_d_a[molecular_idx][2] = np.concatenate(
                            (sum_features_d_a[molecular_idx][2],
                             np.repeat(single_molecule[2][idx, ...][None, ...], repeats=numel, axis=0)),
                            axis=0)
                    except ValueError:
                        sum_features_d_a[molecular_idx][0] = np.repeat(single_molecule[0][idx, ...][None, ...],
                                                                       repeats=numel, axis=0)
                        sum_features_d_a[molecular_idx][1] = np.repeat(single_molecule[1][idx, ...][None, ...],
                                                                       repeats=numel, axis=0)
                        sum_features_d_a[molecular_idx][2] = np.repeat(single_molecule[2][idx, ...][None, ...],
                                                                       repeats=numel, axis=0)

                # Creating SMILES array
                smiles = np.repeat(self.smiles[idx, :][None, :], repeats=numel, axis=0)

                # Generating feature_c
                mask = np.ones(self.features_c_dim, dtype=np.bool)
                mask[selected_c] = 0
                gen_features_c_norm_a = np.zeros((numel, self.features_c_dim))
                # to put into cols of inverse selection, repeat the features_c_a invese columns of selected smiles_idx row
                # for numel number of times to make shape (numel, mask)
                gen_features_c_norm_a[:, mask] = np.repeat(self.features_c_norm_a[idx, mask][None, :],
                                                           repeats=numel, axis=0)

                gen_store = []
                for single_tolerance in selected_tolerance:
                    gen_single_features_c = np.random.rand(numel, 1)
                    if single_tolerance != 0:
                        gen_single_features_c = np.subtract(
                            np.multiply(gen_single_features_c, 1 + single_tolerance * 2), single_tolerance)
                    try:
                        gen_store = np.concatenate((gen_store, gen_single_features_c), axis=1)
                    except np.core._internal.AxisError:
                        gen_store = gen_single_features_c

                gen_features_c_norm_a[:, selected_c] = gen_store
                gen_features_c_a = self.scaler.inverse_transform(gen_features_c_norm_a)

                # sum
                try:
                    sum_features_c_a = np.concatenate((sum_features_c_a, gen_features_c_a), axis=0)
                    sum_smiles = np.concatenate((sum_smiles, smiles), axis=0)
                except ValueError:
                    if sum_features_c_a == []:
                        sum_features_c_a = gen_features_c_a
                    if sum_smiles == []:
                        sum_smiles = smiles
            # Compatibility with mode 0 and 1
            gen_features_c_a = sum_features_c_a
            smiles = sum_smiles
            features_d_a = sum_features_d_a
        elif mode == 3:
            single_dic = {}
            for row_idx, single_example in enumerate(self.smiles.T[1, ...]):
                if single_example not in single_dic:
                    # row_idx is the row number the single_example is in the data_loader.
                    # when comparing row_idx to the excel row number, excel row = row_idx + 2
                    # since excel row counting starts from 1, and excel 1st row is header, so total is plus 2
                    # this is to check if the single_example smiles is not in the dic, add it to the dic along with the idx
                    single_dic[single_example] = row_idx
            keys, values = zip(*single_dic.items())
            smiles_store = list(keys)
            idx_store = list(values)
            numel = len(smiles_store)
            gen_features_c_a = np.repeat(self.features_c_a[smiles_idx, ...][None, :], repeats=numel, axis=0)

            sum_features_d_a = [[[] for _ in range(3)] for _ in range(self.features_d_count)]
            smiles = np.repeat(self.smiles[smiles_idx, ...], repeats=numel, axis=0)

            for count, lig_idx in enumerate(idx_store):
                # Select the molecules for solvent, ligand, RA by choosing smiles_idx
                # First slice for to choose atom, bond, or edge tensor. 2nd slice to choose the specific example data.
                # 3rd slice is to create new axis at 1st dim to repeat numel times to input together with gen_features_c
                for molecular_idx, single_molecule in enumerate(self.features_d_a):
                    if molecular_idx == 0 or molecular_idx == 2:
                        lig_idx = smiles_idx
                    else:
                        lig_idx = lig_idx
                    idx = smiles_idx
                    try:
                        sum_features_d_a[molecular_idx][0] = np.concatenate(
                            (sum_features_d_a[molecular_idx][0], single_molecule[0][idx, ...][None, ...]), axis=0)
                        sum_features_d_a[molecular_idx][1] = np.concatenate(
                            (sum_features_d_a[molecular_idx][1], single_molecule[1][lig_idx, ...][None, ...]), axis=0)
                        sum_features_d_a[molecular_idx][2] = np.concatenate(
                            (sum_features_d_a[molecular_idx][2], single_molecule[2][idx, ...][None, ...]), axis=0)
                    except ValueError:
                        sum_features_d_a[molecular_idx][0] = single_molecule[0][idx, ...][None, ...]
                        sum_features_d_a[molecular_idx][1] = single_molecule[1][lig_idx, ...][None, ...]
                        sum_features_d_a[molecular_idx][2] = single_molecule[2][idx, ...][None, ...]
                smiles[count, 1] = self.smiles[lig_idx, 1]

        else:
            raise TypeError(
                'mode given is {} and is not part of the accepted modes. Try either 0 or 1 as the mode'.format(mode))

        # Create prediction type fl class object
        gen_fl = Features_labels_smiles(gen_features_c_a, features_d_a,
                                        labels=None, smiles=smiles, mode='p', scaler=self.scaler,
                                        features_c_names=self.features_c_names, features_d_names=self.features_d_names,
                                        num_classes=self.n_classes,
                                        save_mode=False)

        return gen_fl

    def write_data_to_excel(self, loader_excel_file):
        # Excel writing part
        wb = load_workbook(loader_excel_file)
        # Setting up subset features and labels sheet

        if self.smiles is not None:
            sheet_name_store = ['temp_features_c', 'temp_features_d', 'temp_labels']
            ss_store = [self.features_c_a, self.smiles, self.labels]
        else:
            sheet_name_store = ['temp_features_c', 'temp_labels']
            ss_store = [self.features_c_a, self.labels]
        axis_store = [2, 0]  # Because feature_c is 2D while labels is col vector, so order of axis is 2,0
        for cnt, sheet_name in enumerate(sheet_name_store):
            if sheet_name in wb.sheetnames:
                # If temp sheet exists, remove it to create a new one. Else skip this.
                idx = wb.sheetnames.index(sheet_name)  # index of temp sheet
                wb.remove(wb.worksheets[idx])  # remove temp
                wb.create_sheet(sheet_name, idx)  # create an empty sheet using old index
            else:
                wb.create_sheet(sheet_name)  # Create the new sheet
            # Print array to the correct worksheet
            print_array_to_excel(ss_store[cnt], (2, 1), wb[sheet_name_store[cnt]], axis=axis_store[cnt])
        wb.save(loader_excel_file)
        wb.close()

    def save(self, save_name):
        file_path = open('./save/features_labels/' + save_name + '.obj', 'wb')
        pickle.dump(self, file_path)
