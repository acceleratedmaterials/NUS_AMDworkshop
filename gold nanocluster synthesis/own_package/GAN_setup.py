import keras as K
from keras.models import Sequential
from keras.layers import Dense, Dropout, Convolution2D, Flatten
from keras import regularizers, optimizers
from keras.callbacks import EarlyStopping, ModelCheckpoint
from keras.layers.advanced_activations import LeakyReLU
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.integrate import odeint


# Own Scripts

def print_array_to_excel(array, first_cell, ws):
    shape = array.shape
    if shape[1] == 1:
        for i in range(shape[0]):
            j = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i][0]
    else:
        for i in range(shape[0]):
            for j in range(shape[1]):
                ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i, j]


def create_GAN_hparams(generator_hidden_layers=[30, 30], discriminator_hidden_layers=[30, 30], learning_rate=None,
                       epochs=100, batch_size=32, activation='relu',
                       optimizer='Adam', loss='mse', patience=4, reg_term=0, generator_dropout=0,
                       discriminator_dropout=0,
                       filters=3):
    """
    Creates hparam dict for input into GAN class. Contain Hyperparameter info
    :return: hparam dict
    """
    names = ['generator_hidden_layers', 'discriminator_hidden_layers', 'learning_rate', 'epochs', 'batch_size',
             'activation', 'optimizer', 'loss', 'patience',
             'reg_term', 'generator_dropout', 'discriminator_dropout',
             'filters']
    values = [generator_hidden_layers, discriminator_hidden_layers, learning_rate, epochs, batch_size, activation,
              optimizer, loss, patience, reg_term, generator_dropout, discriminator_dropout,
              filters]
    hparams = dict(zip(names, values))
    return hparams


def generating_new_examples_using_GAN(numel_new_examples, GAN_generator, x_dim,
                                      gen_examples_dir='./save/gen_examples/',
                                      gen_examples_name='gen_examples.npy',
                                      save_mode=False):
    """
    Uses a previously trained GAN generator model to generate new examples.
    :param numel_new_examples: Number of new examples to generate
    :param features_dim: total number of dimensions = feature nodes (inital,t,T) + label nodes (5 species)
    :param augmented_examples_name: Name of numpy file to save augmented examples to
    :param save_mode: Whether to save the new augmented example in the same folder directory as the original dir
    :return: Augmented examples which contain concatenates the original and GAN generated examples
    """
    z = np.random.normal(0, 1, (numel_new_examples, x_dim))
    v = GAN_generator.predict(z)
    if save_mode:
        # Save excel
        wb = load_workbook('./excel/GAN/gen_examples_template.xlsx')
        ws = wb['gen_examples']
        print_array_to_excel(v, [2, 1], ws)
        wb.save('./excel/GAN/gen_examples_temp.xlsx')
        # Save numpy ndarray
        np.save(gen_examples_dir + gen_examples_name, v)
    return v


def generating_new_examples_using_DAGAN(numel_new_examples, GAN_generator, x_dim=8, real_examples=None,
                                        augmented_examples_name='augmented_training_examples.npy',
                                        real_examples_dir='./save/models/',
                                        real_examples_name='augmented_training_examples.npy',
                                        save_mode=False,
                                        concat_mode=True):
    """
    Uses a previously trained GAN generator model to generate new examples.
    :param numel_new_examples: Number of new examples to generate
    :param features_dim: total number of dimensions = feature nodes (inital,t,T) + label nodes (5 species)
    :param real_examples_dir: Folder directory containing concatenated examples
    :param real_examples_name: Name of numpy file containing the original real examples
    :param augmented_examples_name: Name of numpy file to save augmented examples to
    :param model_dir: Folder directory containing GAN generator model h5 Keras h5 file
    :param model_name: Name of GAN generator model h5 file
    :param save_mode: Whether to save the new augmented example in the same folder directory as the original dir
    :return: Augmented examples which contain concatenates the original and GAN generated examples
    """

    z = np.random.normal(0, 1, (numel_new_examples, features_dim))
    x = np.load(real_examples_dir + real_examples_name)
    batch_x = x[0:numel_new_examples, :]
    batch_x_z = np.concatenate((z, batch_x), axis=1)
    v = GAN_generator.predict(batch_x_z)
    if concat_mode:
        x_v = np.concatenate((x, v), axis=0)
        if save_mode:
            np.save(real_examples_dir + augmented_examples_name, x_v)
        return x_v
    else:
        return v


class GAN:
    def __init__(self, x_dim, z_dim, GAN_hparams):
        """
        Initialise GAN class. Sets up G, D, and the stacked G D model according to GAN_hparams given
        :param features_dim: Generator input dimensions = initial species + 2 (t,T) + 5 (5 output species)
        :param labels_dim: Generator output dimensions = initial species + 2 (t,T) + 5 (5 output species)
        :param GAN_hparams: Dictionary created from create_GAN_hparams function
        """
        self.x_dim = x_dim
        self.z_dim = z_dim
        self.hparams = GAN_hparams

        if GAN_hparams['learning_rate'] is None:
            self.G = self.generator()
            self.G.compile(loss=self.hparams['loss'], optimizer=self.hparams['optimizer'])

            self.D = self.discriminator()
            self.D.compile(loss='binary_crossentropy', optimizer=self.hparams['optimizer'], metrics=['accuracy'])

            self.stacked_generator_discriminator = self.stacked_generator_discriminator()
            self.stacked_generator_discriminator.compile(loss='binary_crossentropy',
                                                         optimizer=self.hparams['optimizer'])
        else:
            sgd = optimizers.Adam(lr=GAN_hparams['learning_rate'])
            self.G = self.generator()
            self.G.compile(loss=self.hparams['loss'], optimizer=sgd)

            self.D = self.discriminator()
            self.D.compile(loss='binary_crossentropy', optimizer=sgd, metrics=['accuracy'])

            self.stacked_generator_discriminator = self.stacked_generator_discriminator()
            self.stacked_generator_discriminator.compile(loss='binary_crossentropy',
                                                         optimizer=sgd)

    def generator(self):
        # Set up Generator model
        generator_input_dim = self.z_dim
        model = Sequential()
        generator_hidden_layers = self.hparams['generator_hidden_layers']
        generator_dropout = self.hparams['generator_dropout']
        if self.hparams['activation'] == 'leakyrelu':
            model.add(Dense(generator_hidden_layers[0],
                            input_dim=generator_input_dim,
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(LeakyReLU(alpha=0.2))
            if generator_dropout != 0:
                model.add(Dropout(generator_dropout))

            numel = len(generator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(generator_hidden_layers[i + 1],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
                    model.add(LeakyReLU(alpha=0.2))
            model.add(Dense(self.x_dim, activation='linear'))
        else:
            model.add(Dense(generator_hidden_layers[0],
                            input_dim=generator_input_dim,
                            activation=self.hparams['activation'],
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            if generator_dropout != 0:
                model.add(Dropout(generator_dropout))

            numel = len(generator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(generator_hidden_layers[i + 1],
                                    activation=self.hparams['activation'],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(Dense(self.x_dim, activation='linear'))

        return model

    def discriminator(self):
        # Set up Discriminator model
        discriminator_input_dim = self.x_dim
        model = Sequential()
        discriminator_hidden_layers = self.hparams['discriminator_hidden_layers']
        discriminator_dropout = self.hparams['discriminator_dropout']

        if self.hparams['activation'] == 'leakyrelu':
            model.add(Dense(discriminator_hidden_layers[0],
                            input_dim=discriminator_input_dim,
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(LeakyReLU(alpha=0.2))
            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            numel = len(discriminator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(discriminator_hidden_layers[i + 1],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
                    model.add(LeakyReLU(alpha=0.2))
            model.add(Dense(1, activation='sigmoid'))
        else:
            model.add(Dense(discriminator_hidden_layers[0],
                            input_dim=discriminator_input_dim,
                            activation=self.hparams['activation'],
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            numel = len(discriminator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(discriminator_hidden_layers[i + 1],
                                    activation=self.hparams['activation'],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(Dense(1, activation='sigmoid'))

        return model

    def stacked_generator_discriminator(self):
        # Freeze discriminator weights and biases when training generator.
        self.D.trainable = False

        model = Sequential()
        model.add(self.G)
        model.add(self.D)

        return model

    def train_GAN(self, training_x, save_name='GAN_generator.h5', save_dir='./save/models/', save_mode=False,
                  plot_mode=False, show_plot=False):
        epochs = self.hparams['epochs']
        batch_size = self.hparams['batch_size']
        numel_rows = training_x.shape[0]
        d_loss_store = []
        g_loss_store = []
        plt.figure()
        plt.title('model loss / acc , (G,D) = (' + str(self.hparams['generator_hidden_layers'][0]) + ',' + str(
            self.hparams['discriminator_hidden_layers'][0]) + ')')
        plt.ylabel('loss / acc')
        plt.xlabel('epoch')

        for cnt in range(epochs):  # Epochs is more like number of steps here. 1 step ==> 1 gradient update
            # Training Discriminator
            # Half batch size for discriminator, since half real half fake data =>combine
            d_batch_size = int(batch_size / 2)
            idx = np.random.randint(0, numel_rows - d_batch_size)  # Index to start drawing x batch_x from training_x
            batch_x = training_x[idx:(idx + d_batch_size), :]  # Correct x
            batch_z = np.random.normal(0, 1, (d_batch_size, self.z_dim))  # Random noise z to feed into G
            batch_v = self.G.predict(batch_z)  # v = f(z), dim_v = dim_x

            combined_x_v = np.concatenate((batch_x, batch_v), axis=0)
            combined_y = np.concatenate((np.ones((d_batch_size, 1)), np.zeros((d_batch_size, 1))), axis=0)

            d_loss = self.D.train_on_batch(combined_x_v, combined_y)  # Returns loss and accuracy
            d_loss_store.append(d_loss)

            # Training Generator using stacked generator, discriminator model
            batch_z = np.random.normal(0, 1, (batch_size, self.z_dim))  # Now is full batch size, not halved
            mislabelled_y = np.ones((batch_size, 1))  # y output all labelled as 1 so that G will train towards that

            g_loss = self.stacked_generator_discriminator.train_on_batch(batch_z, mislabelled_y)
            g_loss_store.append(g_loss)
            if cnt % 10 == 0 or cnt + 1 == epochs:
                print('epoch: %d, [Discriminator :: d_loss: %f , d_acc: %f], [ Generator :: loss: %f]' % (
                    cnt + 1, d_loss[0], d_loss[1], g_loss))

        # Plotting
        if plot_mode:
            d_loss_store = np.array(d_loss_store)
            g_loss_store = np.array(g_loss_store)
            plt.plot(d_loss_store[:, 0])
            plt.plot(d_loss_store[:, 1])
            plt.plot(g_loss_store)
            plt.legend(['d_loss', 'd_acc', 'g_loss'], loc='upper left')
            plt.savefig('./plots/' + str(self.hparams['generator_hidden_layers'][0]) + '_' + str(
                self.hparams['discriminator_hidden_layers'][0]), bbox_inches='tight')
            if show_plot:
                plt.show()
            plt.clf()

        plt.close()

        # Saving
        if save_mode:
            self.G.save(save_dir + save_name)


class DAGAN:
    def __init__(self, augmentation_dim, features_dim, labels_dim, GAN_hparams):
        """
        Initialise GAN class. Sets up G, D, and the stacked G D model according to GAN_hparams given
        :param augmentation_dim: Generator input dimensions = dimension of augmentation z to concatenate with original x
        :param features_dim: Generator input dimensions = initial species + 2 (t,T) + 5 (5 output species)
        :param labels_dim: Generator output dimensions = initial species + 2 (t,T) + 5 (5 output species)
        :param GAN_hparams: Dictionary created from create_GAN_hparams function
        """
        self.augmentation_dim = augmentation_dim
        self.features_dim = features_dim
        self.labels_dim = labels_dim
        self.hparams = GAN_hparams

        if GAN_hparams['learning_rate'] is None:
            self.G = self.generator()
            self.G.compile(loss=self.hparams['loss'], optimizer=self.hparams['optimizer'])

            self.D = self.discriminator()
            self.D.compile(loss='binary_crossentropy', optimizer=self.hparams['optimizer'], metrics=['accuracy'])

            self.stacked_generator_discriminator = self.stacked_generator_discriminator()
            self.stacked_generator_discriminator.compile(loss='binary_crossentropy',
                                                         optimizer=self.hparams['optimizer'])
        else:
            sgd = optimizers.Adam(lr=GAN_hparams['learning_rate'])
            self.G = self.generator()
            self.G.compile(loss=self.hparams['loss'], optimizer=sgd)

            self.D = self.discriminator()
            self.D.compile(loss='binary_crossentropy', optimizer=sgd, metrics=['accuracy'])

            self.stacked_generator_discriminator = self.stacked_generator_discriminator()
            self.stacked_generator_discriminator.compile(loss='binary_crossentropy',
                                                         optimizer=sgd)

    def generator(self):
        # Set up Generator model
        generator_input_dim = self.features_dim + self.augmentation_dim  # Main difference between GAN and DAGAN
        model = Sequential()
        generator_hidden_layers = self.hparams['generator_hidden_layers']
        generator_dropout = self.hparams['generator_dropout']
        if self.hparams['activation'] == 'leakyrelu':
            model.add(Dense(generator_hidden_layers[0],
                            input_dim=generator_input_dim,
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(LeakyReLU(alpha=0.2))
            if generator_dropout != 0:
                model.add(Dropout(generator_dropout))

            numel = len(generator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(generator_hidden_layers[i + 1],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
                    model.add(LeakyReLU(alpha=0.2))
            model.add(Dense(self.labels_dim, activation='linear'))
        else:
            model.add(Dense(generator_hidden_layers[0],
                            input_dim=generator_input_dim,
                            activation=self.hparams['activation'],
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            if generator_dropout != 0:
                model.add(Dropout(generator_dropout))

            numel = len(generator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(generator_hidden_layers[i + 1],
                                    activation=self.hparams['activation'],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(Dense(self.labels_dim, activation='linear'))

        return model

    def discriminator(self):
        # Set up Discriminator model
        discriminator_input_dim = self.features_dim
        model = Sequential()
        discriminator_hidden_layers = self.hparams['discriminator_hidden_layers']
        discriminator_dropout = self.hparams['discriminator_dropout']

        if self.hparams['activation'] == 'leakyrelu':
            model.add(Dense(discriminator_hidden_layers[0],
                            input_dim=discriminator_input_dim,
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(LeakyReLU(alpha=0.2))
            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            numel = len(discriminator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(discriminator_hidden_layers[i + 1],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
                    model.add(LeakyReLU(alpha=0.2))
            model.add(Dense(1, activation='sigmoid'))
        else:
            model.add(Dense(discriminator_hidden_layers[0],
                            input_dim=discriminator_input_dim,
                            activation=self.hparams['activation'],
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            numel = len(discriminator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(discriminator_hidden_layers[i + 1],
                                    activation=self.hparams['activation'],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(Dense(1, activation='sigmoid'))

        return model

    def stacked_generator_discriminator(self):
        # Freeze discriminator weights and biases when training generator.
        self.D.trainable = False

        model = Sequential()
        model.add(self.G)
        model.add(self.D)

        return model

    def train_GAN(self, training_x, save_name='DAGAN_generator.h5', save_dir='./save/models/', save_mode=False,
                  plot_mode=False, show_plot=False):
        epochs = self.hparams['epochs']
        batch_size = self.hparams['batch_size']
        numel_rows = training_x.shape[0]
        d_loss_store = []
        g_loss_store = []
        plt.figure()
        plt.title('model loss / acc , (G,D) = (' + str(self.hparams['generator_hidden_layers'][0]) + ',' + str(
            self.hparams['discriminator_hidden_layers'][0]) + ')')
        plt.ylabel('loss / acc')
        plt.xlabel('epoch')

        for cnt in range(epochs):  # Epochs is more like number of steps here. 1 step ==> 1 gradient update
            # Training Discriminator
            # Half batch size for discriminator, since half real half fake data =>combine
            d_batch_size = int(batch_size / 2)
            idx = np.random.randint(0, numel_rows - d_batch_size)  # Index to start drawing batch_x from training_x
            batch_x = training_x[idx:(idx + d_batch_size), :]  # Correct x
            batch_z = np.random.normal(0, 1, (d_batch_size, self.augmentation_dim))  # Random noise z to feed into G
            batch_z_x = np.concatenate((batch_z, batch_x), axis=1)  # Concat z,x along col to make higher dim input to G
            batch_v = self.G.predict(batch_z_x)  # v = f(z,x)

            combined_x_v = np.concatenate((batch_x, batch_v), axis=0)
            combined_y = np.concatenate((np.ones((d_batch_size, 1)), np.zeros((d_batch_size, 1))), axis=0)

            d_loss = self.D.train_on_batch(combined_x_v, combined_y)  # Returns loss and accuracy
            d_loss_store.append(d_loss)

            # Training Generator using stacked generator, discriminator model
            idx = np.random.randint(0, numel_rows - batch_size)  # Index to start drawing batch_x from training_x
            batch_x = training_x[idx:(idx + batch_size), :]  # Correct x
            batch_z = np.random.normal(0, 1, (batch_size, self.augmentation_dim))  # Now is full batch size, not halved
            batch_z_x = np.concatenate((batch_z, batch_x), axis=1)  # Concat z,x along col to make higher dim input to G
            mislabelled_y = np.ones((batch_size, 1))  # y output all labelled as 1 so that G will train towards that

            g_loss = self.stacked_generator_discriminator.train_on_batch(batch_z_x, mislabelled_y)
            g_loss_store.append(g_loss)
            if cnt % 10 == 0 or cnt + 1 == epochs:
                print('epoch: %d, [Discriminator :: d_loss: %f , d_acc: %f], [ Generator :: loss: %f]' % (
                    cnt + 1, d_loss[0], d_loss[1], g_loss))

        # Plotting
        if plot_mode:
            d_loss_store = np.array(d_loss_store)
            g_loss_store = np.array(g_loss_store)
            plt.plot(d_loss_store[:, 0])
            plt.plot(d_loss_store[:, 1])
            plt.plot(g_loss_store)
            plt.legend(['d_loss', 'd_acc', 'g_loss'], loc='upper left')
            plt.savefig('./plots/' + str(self.hparams['generator_hidden_layers'][0]) + '_' + str(
                self.hparams['discriminator_hidden_layers'][0]), bbox_inches='tight')
            if show_plot:
                plt.show()
            plt.clf()

        plt.close()

        # Saving
        if save_mode:
            self.G.save(save_dir + save_name)


def examples_to_CNN1(input_examples):
    numel = input_examples.shape[0]  # Number of examples
    features = input_examples[:, 0:3]  # Take features which is Ca0, t, T
    labels = input_examples[:, 3:]  # Take labels which is Ca, Cr, Cs, Ct, Cu
    examples = []
    for i in range(numel):
        feature = np.transpose(features[i, None, :])  # Take row vector of 1 example of Ca0, t, T, then transpose to 3x1
        feature = np.repeat(feature, repeats=5, axis=1)  # Repeat 3x1 col vector to form 5 identical cols
        label = labels[i, None, :]  # Take row vector of 1 example of the 5 output species concentration
        example = np.concatenate((feature, label), axis=0)  # Combine the repeated features and 1 row of column
        examples.append(example)  # Add it to list store
    examples = np.array(examples)  # Change multi-dim list to multi-dim np array
    examples = np.reshape(examples, (numel, 1, 4, 5))  # Reshape to input into Keras CNN layer
    return examples


class CGAN:
    def __init__(self, GAN_hparams):
        """
        Initialise GAN class. Sets up G, D, and the stacked G D model according to GAN_hparams given
        For denbigh reaction with 1 initial species and 5 output species only
        :param features_dim: Generator input dimensions = initial species + 2 (t,T) + 5 (5 output species)
        :param labels_dim: Generator output dimensions = initial species + 2 (t,T) + 5 (5 output species)
        :param GAN_hparams: Dictionary created from create_GAN_hparams function
        """
        self.features_dim = features_dim
        self.labels_dim = labels_dim
        self.hparams = GAN_hparams

        if GAN_hparams['learning_rate'] is None:
            self.G = self.generator()
            self.G.compile(loss=self.hparams['loss'], optimizer=self.hparams['optimizer'])

            self.D = self.discriminator()
            self.D.compile(loss='binary_crossentropy', optimizer=self.hparams['optimizer'], metrics=['accuracy'])

            self.stacked_generator_discriminator = self.stacked_generator_discriminator()
            self.stacked_generator_discriminator.compile(loss='binary_crossentropy',
                                                         optimizer=self.hparams['optimizer'])
        else:
            sgd = optimizers.Adam(lr=GAN_hparams['learning_rate'])
            self.G = self.generator()
            self.G.compile(loss=self.hparams['loss'], optimizer=sgd)

            self.D = self.discriminator()
            self.D.compile(loss='binary_crossentropy', optimizer=sgd, metrics=['accuracy'])

            self.stacked_generator_discriminator = self.stacked_generator_discriminator()
            self.stacked_generator_discriminator.compile(loss='binary_crossentropy',
                                                         optimizer=sgd)

    def generator(self):
        # Set up Generator model
        generator_input_dim = self.features_dim
        model = Sequential()
        generator_hidden_layers = self.hparams['generator_hidden_layers']
        generator_dropout = self.hparams['generator_dropout']
        if self.hparams['activation'] == 'leakyrelu':
            model.add(Dense(generator_hidden_layers[0],
                            input_dim=generator_input_dim,
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(LeakyReLU(alpha=0.2))
            if generator_dropout != 0:
                model.add(Dropout(generator_dropout))

            numel = len(generator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(generator_hidden_layers[i + 1],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
                    model.add(LeakyReLU(alpha=0.2))
            model.add(Dense(self.labels_dim, activation='linear'))
        else:
            model.add(Dense(generator_hidden_layers[0],
                            input_dim=generator_input_dim,
                            activation=self.hparams['activation'],
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            if generator_dropout != 0:
                model.add(Dropout(generator_dropout))

            numel = len(generator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(generator_hidden_layers[i + 1],
                                    activation=self.hparams['activation'],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(Dense(self.labels_dim, activation='linear'))

        return model

    def discriminator(self):
        # Set up Discriminator model
        discriminator_input_dim = self.features_dim
        model = Sequential()
        discriminator_hidden_layers = self.hparams['discriminator_hidden_layers']
        discriminator_dropout = self.hparams['discriminator_dropout']

        if self.hparams['activation'] == 'leakyrelu':
            # Convolutional layer
            model.add(Convolution2D(filters=self.hparams['filters'], kernel_size=(4, 1), strides=(1, 1),
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(LeakyReLU(alpha=0.2))
            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            # Dense layers
            model.add(Flatten())
            model.add(Dense(discriminator_hidden_layers[0],
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(LeakyReLU(alpha=0.2))

            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            numel = len(discriminator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(discriminator_hidden_layers[i + 1],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
                    model.add(LeakyReLU(alpha=0.2))
            model.add(Dense(1, activation='sigmoid'))
        else:
            model.add(Convolution2D(filters=self.hparams['filters'], kernel_size=(4, 1), strides=(1, 1),
                                    activation=self.hparams['activation'],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            # Dense layers
            model.add(Flatten())
            model.add(Dense(discriminator_hidden_layers[0], activation=self.hparams['activation'],
                            kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            if discriminator_dropout != 0:
                model.add(Dropout(discriminator_dropout))

            numel = len(discriminator_hidden_layers)
            if numel > 1:
                for i in range(numel - 1):
                    model.add(Dense(discriminator_hidden_layers[i + 1],
                                    activation=self.hparams['activation'],
                                    kernel_regularizer=regularizers.l2(self.hparams['reg_term'])))
            model.add(Dense(1, activation='sigmoid'))

        return model

    def stacked_generator_discriminator(self):
        # Freeze discriminator weights and biases when training generator.
        self.D.trainable = False

        model = Sequential()
        model.add(self.G)
        model.add(self.D)

        return model

    def train_GAN(self, training_x, save_name='GAN_generator.h5', save_dir='./save/models/', save_mode=False,
                  plot_mode=False, show_plot=False):
        epochs = self.hparams['epochs']
        batch_size = self.hparams['batch_size']
        numel_rows = training_x.shape[0]
        d_loss_store = []
        g_loss_store = []
        plt.figure()
        plt.title('model loss / acc , (G,D) = (' + str(self.hparams['generator_hidden_layers'][0]) + ',' + str(
            self.hparams['discriminator_hidden_layers'][0]) + ')')
        plt.ylabel('loss / acc')
        plt.xlabel('epoch')

        for cnt in range(epochs):  # Epochs is more like number of steps here. 1 step ==> 1 gradient update
            # Training Discriminator
            # Half batch size for discriminator, since half real half fake data =>combine
            d_batch_size = int(batch_size / 2)
            idx = np.random.randint(0, numel_rows - d_batch_size)  # Index to start drawing x batch_x from training_x
            batch_x = training_x[idx:(idx + d_batch_size), :]  # Correct x
            batch_z = np.random.normal(0, 1, (d_batch_size, self.features_dim))  # Random noise z to feed into G
            batch_v = self.G.predict(batch_z)  # v = f(z)

            combined_x_v = np.concatenate((batch_x, batch_v), axis=0)
            combined_y = np.concatenate((np.ones((d_batch_size, 1)), np.zeros((d_batch_size, 1))), axis=0)

            d_loss = self.D.train_on_batch(examples_to_CNN1(combined_x_v), combined_y)  # Returns loss and accuracy
            d_loss_store.append(d_loss)

            # Training Generator using stacked generator, discriminator model
            batch_z = np.random.normal(0, 1, (batch_size, self.features_dim))  # Now is full batch size, not halved
            mislabelled_y = np.ones((batch_size, 1))  # y output all labelled as 1 so that G will train towards that

            g_loss = self.stacked_generator_discriminator.train_on_batch(batch_z, mislabelled_y)
            g_loss_store.append(g_loss)
            if cnt % 10 == 0 or cnt + 1 == epochs:
                print('epoch: %d, [Discriminator :: d_loss: %f , d_acc: %f], [ Generator :: loss: %f]' % (
                    cnt + 1, d_loss[0], d_loss[1], g_loss))

        # Plotting
        if plot_mode:
            d_loss_store = np.array(d_loss_store)
            g_loss_store = np.array(g_loss_store)
            plt.plot(d_loss_store[:, 0])
            plt.plot(d_loss_store[:, 1])
            plt.plot(g_loss_store)
            plt.legend(['d_loss', 'd_acc', 'g_loss'], loc='upper left')
            plt.savefig('./plots/' + str(self.hparams['generator_hidden_layers'][0]) + '_' + str(
                self.hparams['discriminator_hidden_layers'][0]), bbox_inches='tight')
            if show_plot:
                plt.show()
            plt.clf()

        plt.close()

        # Saving
        if save_mode:
            self.G.save(save_dir + save_name)
