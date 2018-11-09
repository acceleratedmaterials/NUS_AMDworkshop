import numpy as np
import pandas as pd
from openpyxl import load_workbook
import sys

def print_array_to_excel(array, first_cell, ws, axis=2):
    '''
    Print an np array to excel using openpyxl
    :param array: np array
    :param first_cell: first cell to start dumping values in
    :param ws: worksheet reference. From openpyxl, ws=wb[sheetname]
    :param axis: to determine if the array is a col vector (0), row vector (1), or 2d matrix (2)
    '''
    if isinstance(array, (list,)):
        array = np.array(array)
    shape = array.shape
    if axis == 0:
        # Treat array as col vector and print along the rows
        array.flatten()  # Flatten in case the input array is a nx1 ndarry which acts weird
        for i in range(shape[0]):
            j = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i]
    elif axis == 1:
        # Treat array as row vector and print along the columns
        array.flatten()  # Flatten in case the input array is a 1xn ndarry which acts weird
        for j in range(shape[0]):
            i = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[j]
    elif axis == 2:
        # If axis==2, means it is a 2d array
        for i in range(shape[0]):
            for j in range(shape[1]):
                ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i, j]


if __name__ == '__main__':
    print('hi')