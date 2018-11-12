import tensorflow as tf
import keras.backend as K
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from sklearn.metrics import confusion_matrix, accuracy_score, f1_score, matthews_corrcoef, mean_squared_error
import matplotlib.pyplot as plt
from skopt import gp_minimize
from skopt.space import Real, Categorical, Integer
from skopt.utils import use_named_args
from skopt.plots import plot_convergence
import pickle, time, os, gc
# Own Scripts
from own_package.features_labels_setup import read_reaction_data, read_reaction_data_smiles
from own_package.SNN_setup import create_SNN_hparams, SNN, Siamese_loader, SNN_smiles, Siamese_loader_smiles
from own_package.DNN_setup import DNN_classifer, DNN, DNN_classifer_smiles
from own_package.SVM_DT_setup import SVM_smiles, SVM
from own_package.others import print_array_to_excel


def run_skf(model_mode, cv_mode, hparams, loader_file, skf_file='./excel/skf.xlsx', skf_sheet=None,
            k_folds=10, k_shuffle=True, save_model=False, save_model_name=None, save_model_dir='./save/models/'):
    '''
    Stratified k fold cross validation for training and evaluating model 2 only. Model 1 data is trained before hand.
    :param model_mode: Choose between using SNN or cDNN (non_smiles) and SNN_smiles or cDNN_smiles
    :param cv_mode: Cross validation mode. Either 'skf' or 'loocv'.
    :param hparams: hparams dict containing hyperparameters information
    :param loader_file: data_loader excel file location
    :param skf_file: skf_file name to save excel file as
    :param skf_sheet: name of sheet to save inside the skf_file excel. If None, will default to SNN or cDNN as name
    :param k_folds: Number of k folds. Used only for skf cv_mode
    :param k_shuffle: Whether to shuffle the given examples to split into k folds if using skf
    :return:
    '''
    # Choosing between smiles vs non-smiles
    if model_mode == 'SNN_smiles' or model_mode == 'cDNN_smiles' or model_mode == 'SVM_smiles':
        # Smiles mode
        fl = read_reaction_data_smiles(loader_file, mode='c', save_mode=False)
        smiles_mode = True
    else:
        # Non-smiles mode
        fl = read_reaction_data(loader_file, mode='c', save_mode=False)
        smiles_mode = False

    # Creating k-folds
    if cv_mode == 'skf':
        fl_store = fl.create_kf(k_folds=k_folds, shuffle=k_shuffle)
    elif cv_mode == 'loocv':
        fl_store = fl.create_loocv()
    else:
        raise TypeError('cv_mode should be a string containing either skf or loocv to choose either one.'
                        ' {} was given instead.'.format(cv_mode))

    # Run k model instance to perform skf
    predicted_labels_store = []
    acc_store = []
    ce_store = []
    f1s_store = []
    mcc_store = []
    folds = []
    val_idx = []
    val_features_c = []
    val_smiles = []
    val_labels = []
    for fold, fl_tuple in enumerate(fl_store):
        sess = tf.Session()
        K.set_session(sess)
        instance_start = time.time()
        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl
        if model_mode == 'SNN':
            # Run SNN
            model = SNN(hparams, ss_fl)
            loader = Siamese_loader(model.siamese_net, ss_fl, hparams)
            loader.train(loader.hparams.get('epochs', 100), loader.hparams.get('batch_size', 32),
                         verbose=loader.hparams.get('verbose', 1))
            predicted_labels, acc, ce, cm, f1s, mcc = loader.eval(i_ss_fl)
            predicted_labels_store.extend(predicted_labels)
            acc_store.append(acc)
            ce_store.append(ce)
            f1s_store.append(f1s)
            mcc_store.append(mcc)
            if save_model:
                # Set save_model_name
                if isinstance(save_model_name, str):
                    save_model_name1 = save_model_name + '_' + model_mode + '_' + cv_mode + '_' + str(fold + 1)
                else:
                    save_model_name1 = model_mode + '_' + cv_mode + '_' + str(fold + 1)
                # Checking if save model name file already exists, if so, add word 'new' behind
                if os.path.isfile(save_model_dir + save_model_name1 + '.h5'):
                    save_model_name1 = 'new_' + save_model_name1
                # Save model
                print('Saving instance {} model in {}'.format(fold + 1, save_model_dir + save_model_name1 + '.h5'))
                model.siamese_net.save(save_model_dir + save_model_name1 + '.h5')
            del loader  # Need to put this if not memory will run out
        elif model_mode == 'cDNN' or model_mode == 'SVM':
            # Run DNN
            if model_mode == 'cDNN_smiles':
                model = DNN_classifer(hparams, ss_fl)
            else:
                model = SVM(hparams, ss_fl)
            model.train_model(ss_fl)
            predicted_labels, acc, ce, cm, f1s, mcc = model.eval(i_ss_fl)
            predicted_labels_store.extend(predicted_labels)
            acc_store.append(acc)
            ce_store.append(ce)
            f1s_store.append(f1s)
            mcc_store.append(mcc)
            if save_model:
                # Set save_model_name
                if isinstance(save_model_name, str):
                    save_model_name1 = save_model_name + '_' + model_mode + '_' + cv_mode + '_' + str(fold + 1)
                else:
                    save_model_name1 = model_mode + '_' + cv_mode + '_' + str(fold + 1)
                # Checking if save model name file already exists, if so, add word 'new' behind
                if os.path.isfile(save_model_dir + save_model_name1 + '.h5'):
                    save_model_name1 = 'new_' + save_model_name1
                # Save model
                print('Saving instance {} model in {}'.format(fold + 1, save_model_dir + save_model_name1 + '.h5'))
                model.model.save(save_model_dir + save_model_name1 + '.h5')
        elif model_mode == 'cDNN_smiles' or model_mode == 'SVM_smiles':
            # Run DNN or SVM for smiles. Those two are put together because they only differ in the first line of code.
            if model_mode == 'cDNN_smiles':
                model = DNN_classifer_smiles(hparams, ss_fl)
            else:
                model = SVM_smiles(hparams, ss_fl)
            model.train_model(ss_fl)
            predicted_labels, acc, ce, cm, f1s, mcc = model.eval(i_ss_fl)
            predicted_labels_store.extend(predicted_labels)
            acc_store.append(acc)
            ce_store.append(ce)
            f1s_store.append(f1s)
            mcc_store.append(mcc)
            if save_model:
                # Set save_model_name
                if isinstance(save_model_name, str):
                    save_model_name1 = save_model_name + '_' + model_mode + '_' + cv_mode + '_' + str(fold + 1)
                else:
                    save_model_name1 = model_mode + '_' + cv_mode + '_' + str(fold + 1)
                # Checking if save model name file already exists, if so, add word 'new' behind
                if os.path.isfile(save_model_dir + save_model_name1 + '.h5'):
                    save_model_name1 = 'new_' + save_model_name1
                # Save model
                print('Saving instance {} model in {}'.format(fold + 1, save_model_dir + save_model_name1 + '.h5'))
                model.model.save(save_model_dir + save_model_name1 + '.h5')
        elif model_mode == 'SNN_smiles':
            # Run SNN_smiles
            model = SNN_smiles(hparams, ss_fl)
            loader = Siamese_loader_smiles(model.siamese_net, ss_fl, hparams)
            loader.train(loader.hparams.get('epochs', 100), loader.hparams.get('batch_size', 32),
                         loader.hparams.get('pair_size', 32), verbose=loader.hparams.get('verbose', 1))
            predicted_labels, acc, ce, cm, f1s, mcc = loader.eval(i_ss_fl)
            predicted_labels_store.extend(predicted_labels)
            acc_store.append(acc)
            ce_store.append(ce)
            f1s_store.append(f1s)
            mcc_store.append(mcc)
            if save_model:
                # Set save_model_name
                if isinstance(save_model_name, str):
                    save_model_name1 = save_model_name + '_' + model_mode + '_' + cv_mode + '_' + str(fold + 1)
                else:
                    save_model_name1 = model_mode + '_' + cv_mode + '_' + str(fold + 1)
                # Checking if save model name file already exists, if so, add word 'new' behind
                if os.path.isfile(save_model_dir + save_model_name1 + '.h5'):
                    save_model_name1 = 'new_' + save_model_name1
                # Save model
                print('Saving instance {} model in {}'.format(fold + 1, save_model_dir + save_model_name1 + '.h5'))
                model.siamese_net.save(save_model_dir + save_model_name1 + '.h5')
            del loader  # Need to put this if not memory will run out
        else:
            raise TypeError('model_mode {} is not in the list of acceptable model_mode. Input string of either'
                            'SNN, cDNN, SNN_smiles'.format(model_mode))

        # Need to put the next 3 lines if not memory will run out
        del model
        K.clear_session()
        gc.collect()

        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        val_features_c = np.concatenate((val_features_c, i_ss_fl.features_c_a),
                                        axis=0) if val_features_c != [] else i_ss_fl.features_c_a

        if smiles_mode:
            val_smiles = np.concatenate((val_smiles, i_ss_fl.smiles),
                                        axis=0) if val_smiles != [] else i_ss_fl.smiles

        val_labels.extend(i_ss_fl.labels)
        val_idx.extend(i_ss_fl.idx)

        # Printing one instance summary.
        instance_end = time.time()
        if cv_mode == 'skf':
            print(
                '\nFor k-fold run {} out of {}. Each fold has {} examples. Model is {}. Time taken for instance = {}\n'
                'Post-training results: \nacc = {} , ce = {} , f1 score = {} , mcc = {}\ncm = \n{}\n'
                '####################################################################################################'
                    .format(fold + 1, k_folds, i_ss_fl.count, model_mode, instance_end - instance_start, acc, ce, f1s,
                            mcc,
                            cm))
        else:
            print('\nFor LOOCV run {} out of {}. Model is {}. Time taken for instance = {}\n'
                  'Post-training results: \nacc = {} , ce = {} , f1 score = {} , mcc = {}\ncm = \n{}\n'
                  '####################################################################################################'
                  .format(fold + 1, fl.count, model_mode, instance_end - instance_start, acc, ce, f1s, mcc, cm))

    acc_avg = np.average(acc_store)
    ce_avg = np.average(ce_store)
    f1s_avg = np.average(f1s_store)
    f1s_var = np.var(f1s_store)
    mcc_avg = np.average(mcc_store)
    mcc_var = np.var(mcc_store)

    # Creating dataframe to print into excel later.
    if smiles_mode:
        new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                                 val_features_c,
                                 val_smiles,
                                 np.array(val_labels)[:, None],
                                 np.array(predicted_labels_store)[:, None])
                                , axis=1)
        headers = ['folds'] + \
                  ['f' + str(+idx + 1) for idx in range(fl.features_c_count)] + \
                  ['d' + str(+idx + 1) for idx in range(fl.features_d_count)] + \
                  ['Class'] + \
                  ['P_Class']
    else:
        new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                                 val_features_c,
                                 np.array(val_labels)[:, None],
                                 np.array(predicted_labels_store)[:, None])
                                , axis=1)
        headers = ['folds'] + \
                  ['f' + str(+idx + 1) for idx in range(fl.features_c_count)] + \
                  ['Class'] + \
                  ['P_Class']

    # val_idx is the original position of the example in the data_loader
    new_df = pd.DataFrame(data=new_df, columns=headers, index=val_idx)

    # Calculating metrics based on complete validation prediction
    acc_full = accuracy_score(val_labels, predicted_labels_store)
    f1s_full = f1_score(val_labels, predicted_labels_store)
    mcc_full = matthews_corrcoef(val_labels, predicted_labels_store)
    cm_full = confusion_matrix(val_labels, predicted_labels_store)

    # Checking if skf_file excel exists. If not, create new excel
    if skf_file[-5:] != '.xlsx':  # In case you forgotten to put a .xlsx at the back of the excel file string
        skf_file = skf_file + '.xlsx'
    if os.path.isfile(skf_file) and os.access(skf_file, os.W_OK):  # Check if file exists and if file is write-able
        print('Writing into' + skf_file)
        wb = load_workbook(skf_file)
    elif cv_mode == 'skf':
        # Check if the skf_file name is a proper excel file extension, if not, add .xlsx at the back
        print('skf_file not found. Creating new skf_file named as : ' + skf_file)
        wb = openpyxl.Workbook()
        wb.save(skf_file)
    elif cv_mode == 'loocv':
        # Check if the skf_file name is a proper excel file extension, if not, add .xlsx at the back
        # Replace skf with loocv
        print('loocv_file not found. Creating new loocv_file named as : ' + skf_file)
        wb = openpyxl.Workbook()
        wb.save(skf_file)

    # Creating new worksheet. Even if SNN worksheet already exists, a new SNN1 ws will be created and so on
    if skf_sheet is None:
        wb.create_sheet(model_mode)
    else:
        wb.create_sheet(model_mode + skf_sheet)
    sheet_name = wb.sheetnames[-1]  # Taking the ws name from the back ensures that if SNN1 is the new ws, it works

    # Writing hparam dataframe first
    pd_writer = pd.ExcelWriter(skf_file, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    new_df.to_excel(pd_writer, sheet_name)
    start_col = len(new_df.columns) + 3
    hparams = pd.DataFrame(hparams)
    hparams.to_excel(pd_writer, sheet_name, startrow=0, startcol=start_col - 1)
    start_row = 5

    # Writing other subset split, instance per run, and bounds
    ws = wb[sheet_name]
    headers = ['acc', 'ce', 'f1', 'f1_var', 'mcc', 'mcc_var']
    values = [acc_avg, ce_avg, f1s_avg, f1s_var, mcc_avg, mcc_var]
    values_full = [acc_full, -1, f1s_full, -1, mcc_full, -1]
    print_array_to_excel(np.array(headers), (1 + start_row, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values), (2 + start_row, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values_full), (3 + start_row, start_col + 1), ws, axis=1)
    ws.cell(2 + start_row, start_col).value = 'Folds avg'
    ws.cell(3 + start_row, start_col).value = 'Overall'
    ws.cell(4 + start_row, start_col).value = 'Overall cm'
    print_array_to_excel(np.array(cm_full), (4 + start_row, start_col + 1), ws, axis=2)
    if cv_mode == 'skf':
        ws.cell(1, start_col).value = 'SKF'
    elif cv_mode == 'loocv':
        ws.cell(1, start_col).value = 'LOOCV'
    ws.cell(1, start_col - 1).value = loader_file
    pd_writer.save()
    pd_writer.close()
    wb.close()

    return mcc_full


def hparam_opt(model_mode, loader_file, total_run, instance_per_run=3, hparam_file='./excel/hparams_opt.xlsx'):
    names = ['hl_1_l', 'hl_1_h',
             'hl_2_l', 'hl_2_h',
             'fv_l', 'fv_h',
             'epochs_l', 'epochs_h',
             'l1l2_l', 'l1l2_h'
             ]
    values = [5, 500,
              0, 500,
              1, 500,
              20, 200,
              0, 0.1]
    bounds = dict(zip(names, values))
    hl_1 = Integer(low=bounds['hl_1_l'], high=bounds['hl_1_h'], name='hidden_layer_1')
    hl_2 = Integer(low=bounds['hl_2_l'], high=bounds['hl_2_h'], name='hidden_layer_2')
    epochs = Integer(low=bounds['epochs_l'], high=bounds['epochs_h'], name='epochs')
    feature_vector_dim = Integer(low=bounds['fv_l'], high=bounds['fv_h'], name='feature_vector_dim')
    l1l2 = Real(low=bounds['l1l2_l'], high=bounds['l1l2_h'], name='l1l2')
    dimensions = [hl_1, hl_2, feature_vector_dim, epochs, l1l2]
    default_parameters = [20, 0, 10, 25, 0.001]

    global run_count, best_loss, data_store, fl, best_hparams
    run_count = 0
    best_loss = -1000  # Best loss is now best MCC. Lower MCC means worse performance

    @use_named_args(dimensions=dimensions)
    def fitness(hidden_layer_1, hidden_layer_2, feature_vector_dim, epochs, l1l2):
        global run_count, best_loss, data_store, fl, best_hparams
        run_count += 1

        hparams = create_SNN_hparams(epochs=epochs, batch_size=32, pair_size=32,
                                     hidden_layers=[hidden_layer_1, hidden_layer_2],
                                     feature_vector_dim=feature_vector_dim,
                                     dropout=0.5, singlenet_l1=l1l2, singlenet_l2=l1l2, verbose=0,
                                     fp_length=30, fp_number=3, conv_width=32, conv_number=2,
                                     conv_activation='leakyrelu',
                                     conv_l1=0.001, conv_l2=0.001)

        mcc_avg = 0

        for cnt in range(instance_per_run):
            mcc = run_skf(model_mode=model_mode, cv_mode='skf', hparams=hparams,
                          loader_file=loader_file,
                          skf_file=hparam_file, skf_sheet=str(model_mode) + '_' + str(run_count) + '_' + str(cnt),
                          k_folds=10, k_shuffle=True,
                          save_model_name='new_'+str(model_mode) + '_' + str(run_count) + '_' + str(cnt), save_model=True,
                          save_model_dir='./save/models')
            mcc_avg += mcc

        mcc_avg = mcc_avg / instance_per_run
        if mcc_avg > best_loss:
            best_hparams = hparams
            best_loss = mcc_avg
        loss = -mcc_avg  # Negative because we want to maximise mcc, so negative will become minimize
        print('**************************************************************************************************\n'
              'Run Number {} \n'
              'Instance per run {} \n'
              'Current run MCC {} \n'
              '*********************************************************************************************'.format(
            run_count, instance_per_run, mcc_avg))
        print(pd.DataFrame(hparams))
        return loss

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)
    plot_convergence(search_result)
    print('Best Loss = {}'.format(search_result.fun))
    print('Best hparams :')
    best_hparams = pd.DataFrame(best_hparams)
    print(best_hparams)
