import keras.backend as K
import tensorflow as tf
from keras.models import load_model
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os, time, gc

# Own Scripts
from own_package.features_labels_setup import read_reaction_data_smiles
from own_package.SNN_setup import Siamese_loader_smiles, Siamese_loader
from own_package.DNN_setup import DNN_classifer_smiles
from own_package.others import print_array_to_excel
# NGF
from own_package.NGF.layers import NeuralGraphHidden, NeuralGraphOutput


def generate_examples(model_mode, smiles_idx, numel, gen_mode, model_directory, gen_excel_name, selected_c=None,
                      selected_range=None, selected_tolerance=None, gen_sheet_name=None,
                      loader_excel_file='./excel/data_loader/gold_data_loader',
                      gen_excel_dir='./excel/gen/'):
    fl = read_reaction_data_smiles(loader_excel_file=loader_excel_file)
#    fl, eval_fl = fl.create_subsetsplit([-1])     ##I dont think this line make sense
    eval_fl = fl.generate_examples(smiles_idx=smiles_idx, numel=numel, mode=gen_mode,
                                   selected_c=selected_c,
                                   selected_tolerance=selected_tolerance,
                                   selected_range=selected_range)
    model_store = []
    directory = model_directory
    for idx, file in enumerate(os.listdir(directory)):
        filename = os.fsdecode(file)
#        filename = file.split('/')[-1]
        if filename.endswith(".h5"):
            model_store.append(directory + '/' + filename)
        else:
            continue
    print('Loading the following {} models from {}. Total models = {}'.format(model_mode, directory, len(model_store)))

    if model_mode == 'SNN_smiles':
        smiles_mode=True
        loader = Siamese_loader_smiles(model_store, fl, None)
        # (no. of examples, no. of class) , (no, of examples, no. of models) , (no. of examples, no. of class)
        ensemble_predicted_class_store, final_ensemble_predicted_class, final_ensemble_predicted_class_raw \
            = loader.eval_ensemble(eval_fl)
    elif model_mode == 'cDNN_smiles' or model_mode == 'SVM_smiles':
        smiles_mode = True
        # function to evaluated keras model. Since there is no cDNN loader, I put this function here
        def eval(eval_fl, model):
            eval_start = time.time()
            features_c_norm_a = eval_fl.features_c_norm_a
            features_d_a = eval_fl.features_d_a
            features = [[] for _ in range(1 + eval_fl.features_d_count * 3)]
            # Append features c to the 0th list in the pairs nested list
            idx = 0
            features[idx] = features_c_norm_a
            # Add one to the counter idx that is keeping track of the current position in the nested list
            idx += 1
            # Adding information about features_d
            for single_molecule in features_d_a:
                for single_tensor in single_molecule:
                    features[idx] = single_tensor
                    idx += 1
            predictions = model.predict(features)
            predictions_class = np.argmax(predictions, axis=1)

            eval_end = time.time()
            print('eval run time : {}'.format(eval_end - eval_start))
            return predictions_class, predictions

        total_model_count = len(model_store)
        ensemble_predicted_class_store = []
        ensemble_labels_hot_store = []

        for instance, model_path in enumerate(model_store):
            # Stuff to ensure keras can run smoothly due to training multiple model instances in the same script
            sess = tf.Session()
            K.set_session(sess)

            eval_start = time.time()
            model = load_model(model_path, custom_objects={'NeuralGraphHidden': NeuralGraphHidden,
                                                           'NeuralGraphOutput': NeuralGraphOutput})

            predicted_class, predicted_labels_hot = eval(eval_fl, model)
            ensemble_predicted_class_store.append(predicted_class)  # shape (No. of models, n_examples)
            if ensemble_labels_hot_store != []:
                ensemble_labels_hot_store += predicted_labels_hot
            else:
                ensemble_labels_hot_store = predicted_labels_hot

            # Stuff to ensure keras can run smoothly due to training multiple model instances in the same script
            del model
            K.clear_session()
            gc.collect()
            eval_end = time.time()

            print('{} out of {}: {} run time  {}'.format(instance + 1, total_model_count, model_path,
                                                         eval_end - eval_start))

        final_ensemble_predicted_class = np.argmax(ensemble_labels_hot_store, axis=1)
        final_ensemble_predicted_class_raw = np.divide(ensemble_labels_hot_store, total_model_count)

    elif model_mode == 'SNN':
        smiles_mode = False
        loader = Siamese_loader(model_store, fl, None)
        # (no. of examples, no. of class) , (no, of examples, no. of models) , (no. of examples, no. of class)
        ensemble_predicted_class_store, final_ensemble_predicted_class, final_ensemble_predicted_class_raw \
            = loader.eval_ensemble(eval_fl)

    elif model_mode == 'cDNN' or model_mode == 'SVM':
        smiles_mode = False
        # function to evaluated keras model. Since there is no cDNN loader, I put this function here
        def eval(eval_fl, model):
            eval_start = time.time()
            features = eval_fl.features_c_norm_a

            predictions = model.predict(features)
            predictions_class = np.argmax(predictions, axis=1)

            eval_end = time.time()
            print('eval run time : {}'.format(eval_end - eval_start))
            return predictions_class, predictions

        total_model_count = len(model_store)
        ensemble_predicted_class_store = []
        ensemble_labels_hot_store = []

        for instance, model_path in enumerate(model_store):
            # Stuff to ensure keras can run smoothly due to training multiple model instances in the same script
            sess = tf.Session()
            K.set_session(sess)

            eval_start = time.time()
            model = load_model(model_path)

            predicted_class, predicted_labels_hot = eval(eval_fl, model)
            ensemble_predicted_class_store.append(predicted_class)  # shape (No. of models, n_examples)
            if ensemble_labels_hot_store != []:
                ensemble_labels_hot_store += predicted_labels_hot
            else:
                ensemble_labels_hot_store = predicted_labels_hot

            # Stuff to ensure keras can run smoothly due to training multiple model instances in the same script
            del model
            K.clear_session()
            gc.collect()
            eval_end = time.time()

            print('{} out of {}: {} run time  {}'.format(instance + 1, total_model_count, model_path,
                                                         eval_end - eval_start))

        final_ensemble_predicted_class = np.argmax(ensemble_labels_hot_store, axis=1)
        final_ensemble_predicted_class_raw = np.divide(ensemble_labels_hot_store, total_model_count)
    else:
        raise TypeError('model_mode not part of accepted list.')

    # Excel writing part
    # Checking if file names are given correctly
    if gen_excel_dir[-1:] != '/':  # In case you forgotten to add a / at the back of the excel directory string
        gen_excel_dir = gen_excel_dir + '/'
    if gen_excel_name[-5:] != '.xlsx':  # In case you forgotten to put a .xlsx at the back of the excel file string
        gen_excel_name = gen_excel_name + '.xlsx'
    gen_excel_name = gen_excel_dir + gen_excel_name

    # Writing into existing workbook or creating new workbook if workbook not found
    if os.path.isfile(gen_excel_name):
        print('Writing into' + gen_excel_name)
        wb = load_workbook(gen_excel_name)
    else:
        print('gen_excel_file not found. Creating new file named as : ' + gen_excel_name)
        wb = openpyxl.Workbook()
        wb.save(gen_excel_name)

    # Creating new worksheet
    if gen_sheet_name is None:
        wb.create_sheet('gen_' + model_mode)
    else:
        # If special sheet name is given, add sheet name to the back of gen_ string
        wb.create_sheet('gen_' + model_mode + gen_sheet_name)
    sheet_name = wb.sheetnames[-1]  # Taking the ws name from the back to ensure the newly created sheet is selected

    if smiles_mode:
        # Making a df containing all the information
        headers = ['f' + str(+idx + 1) for idx in range(fl.features_c_count)] + \
                  ['d' + str(+idx + 1) for idx in range(fl.features_d_count)] + \
                  ['Ensemble_P'] + \
                  ['E' + str(+idx + 1) for idx in range(final_ensemble_predicted_class_raw.shape[1])] + \
                  ['P' + str(+idx + 1) for idx in range(len(model_store))]
        final_ensemble_predicted_class = final_ensemble_predicted_class[:, None]  # Change 1D to (None,1)
        ensemble_predicted_class_store = np.transpose(ensemble_predicted_class_store)

        df = np.concatenate((eval_fl.features_c_a, eval_fl.smiles, final_ensemble_predicted_class,
                             final_ensemble_predicted_class_raw, ensemble_predicted_class_store), axis=1)
        df = pd.DataFrame(data=df, columns=headers)
    else:
        # Making a df containing all the information
        headers = ['f' + str(+idx + 1) for idx in range(fl.features_c_count)] + \
                  ['Ensemble_P'] + \
                  ['E' + str(+idx + 1) for idx in range(final_ensemble_predicted_class_raw.shape[1])] + \
                  ['P' + str(+idx + 1) for idx in range(len(model_store))]
        final_ensemble_predicted_class = final_ensemble_predicted_class[:, None]  # Change 1D to (None,1)
        ensemble_predicted_class_store = np.transpose(ensemble_predicted_class_store)

        df = np.concatenate((eval_fl.features_c_a, final_ensemble_predicted_class,
                             final_ensemble_predicted_class_raw, ensemble_predicted_class_store), axis=1)
        df = pd.DataFrame(data=df, columns=headers)

    # Print df to excel
    pd_writer = pd.ExcelWriter(gen_excel_name, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    df.to_excel(pd_writer, sheet_name)
    pd_writer.save()
    pd_writer.close()
    wb.close()
    return
'''
generate_examples(model_mode='SNN_smiles', loader_excel_file='./excel/data_loader/gold_data_loader_Au25',
                  smiles_idx=2, numel=5, gen_mode=1, selected_c=[3,4], selected_tolerance=[0.1,0.1,0.1,0.1,0.1],
                  selected_range=[[7, 14], [0, 100]], model_directory='./save/models/final/SNN_smiles_desc_Au25',
                  gen_excel_name='gen')
'''

'''
generate_examples(model_mode='SNN_smiles', loader_excel_file='./excel/data_loader/gold_data_loader_aq_Au25',
                  smiles_idx=2, numel=300, gen_mode=2, selected_c=[0,1,2,3], selected_tolerance=[0.2,0.1,0,0.1],
                  selected_range=[[7, 14], [0, 100]], model_directory='./save/models/final/SNN_smiles_desc_aq_Au25',
                  gen_excel_name='gen_aq_Au25')

'''
